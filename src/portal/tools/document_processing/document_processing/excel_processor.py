"""
Excel Processor Tool
====================

Comprehensive Excel (XLSX) file manipulation.

Features:
- Read/write XLSX files
- Data analysis and statistics
- Formatting and styling
- Formula creation
- Multiple sheets
- Charts and graphs
- Data validation
- Pivot tables (read-only)

Install: pip install openpyxl xlsxwriter pandas
"""

import asyncio
import logging
from typing import Dict, Any, List, Optional, Union
from pathlib import Path

from portal.core.interfaces.tool import BaseTool, ToolMetadata, ToolParameter, ToolCategory

logger = logging.getLogger(__name__)


class ExcelProcessorTool(BaseTool):
    """
    Comprehensive Excel file processor.
    
    Create, read, analyze, and format Excel spreadsheets.
    """
    
    def __init__(self):
        super().__init__()
    
    def _get_metadata(self) -> ToolMetadata:
        return ToolMetadata(
            name="excel_processor",
            description="Create, read, analyze, and format Excel (XLSX) files",
            category=ToolCategory.DATA,
            version="1.0.0",
            requires_confirmation=False,
            parameters=[
                ToolParameter(
                    name="action",
                    param_type="string",
                    description="Action: read, write, analyze, format, add_chart",
                    required=True
                ),
                ToolParameter(
                    name="file_path",
                    param_type="string",
                    description="Path to Excel file",
                    required=True
                ),
                ToolParameter(
                    name="sheet_name",
                    param_type="string",
                    description="Sheet name (default: Sheet1)",
                    required=False,
                    default="Sheet1"
                ),
                ToolParameter(
                    name="data",
                    param_type="object",
                    description="Data to write (list of lists or dict)",
                    required=False
                ),
                ToolParameter(
                    name="range",
                    param_type="string",
                    description="Cell range (e.g., 'A1:C10')",
                    required=False
                ),
                ToolParameter(
                    name="headers",
                    param_type="list",
                    description="Column headers",
                    required=False
                ),
                ToolParameter(
                    name="formatting",
                    param_type="object",
                    description="Formatting options (font, fill, border, alignment)",
                    required=False
                ),
                ToolParameter(
                    name="chart_type",
                    param_type="string",
                    description="Chart type: bar, line, pie",
                    required=False
                ),
                ToolParameter(
                    name="formulas",
                    param_type="list",
                    description="Formulas to add (list of cell-formula pairs)",
                    required=False
                )
            ]
        )
    
    async def execute(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Execute Excel operation"""

        # Lazy import - only load when actually executing
        try:
            import openpyxl
        except ImportError:
            return self._error_response(
                "openpyxl not installed. Install: pip install openpyxl xlsxwriter"
            )

        action = parameters.get("action", "").lower()

        if action == "read":
            return await self._read_excel(parameters)
        elif action == "write":
            return await self._write_excel(parameters)
        elif action == "analyze":
            return await self._analyze_excel(parameters)
        elif action == "format":
            return await self._format_excel(parameters)
        elif action == "add_chart":
            return await self._add_chart(parameters)
        else:
            return self._error_response(f"Unknown action: {action}")
    
    async def _read_excel(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Read Excel file"""

        import openpyxl

        file_path = Path(parameters.get("file_path", "")).expanduser()
        sheet_name = parameters.get("sheet_name", "Sheet1")
        cell_range = parameters.get("range")

        if not file_path.exists():
            return self._error_response(f"File not found: {file_path}")

        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Get sheet
            if sheet_name not in wb.sheetnames:
                return self._error_response(f"Sheet '{sheet_name}' not found")
            
            ws = wb[sheet_name]
            
            # Read data
            if cell_range:
                data = [[cell.value for cell in row] for row in ws[cell_range]]
            else:
                data = [[cell.value for cell in row] for row in ws.iter_rows()]
            
            # Get sheet info
            metadata = {
                "sheets": wb.sheetnames,
                "active_sheet": sheet_name,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "has_formulas": any(
                    cell.data_type == 'f'
                    for row in ws.iter_rows()
                    for cell in row
                )
            }
            
            wb.close()
            
            return self._success_response(
                result={"data": data},
                metadata=metadata
            )
        
        except Exception as e:
            logger.error(f"Excel read error: {e}")
            return self._error_response(f"Read error: {e}")
    
    async def _write_excel(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Write Excel file"""

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        file_path = Path(parameters.get("file_path", "")).expanduser()
        sheet_name = parameters.get("sheet_name", "Sheet1")
        data = parameters.get("data")
        headers = parameters.get("headers")
        formulas = parameters.get("formulas", [])

        if not data:
            return self._error_response("No data provided")

        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Write headers
            start_row = 1
            if headers:
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    # Bold headers
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                start_row = 2
            
            # Write data
            if isinstance(data, dict):
                # Convert dict to rows
                data = list(data.values())
            
            for row_idx, row_data in enumerate(data, start=start_row):
                if isinstance(row_data, dict):
                    row_data = list(row_data.values())
                
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add formulas
            for formula_spec in formulas:
                cell_ref = formula_spec.get("cell")
                formula = formula_spec.get("formula")
                if cell_ref and formula:
                    ws[cell_ref] = formula
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save
            file_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(file_path)
            wb.close()
            
            return self._success_response(
                result={"file_path": str(file_path)},
                metadata={
                    "rows_written": len(data),
                    "columns": len(data[0]) if data else 0,
                    "formulas_added": len(formulas)
                }
            )
        
        except Exception as e:
            logger.error(f"Excel write error: {e}")
            return self._error_response(f"Write error: {e}")
    
    async def _analyze_excel(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze Excel data"""

        try:
            import pandas as pd
        except ImportError:
            return self._error_response("pandas not installed. Install: pip install pandas")

        file_path = Path(parameters.get("file_path", "")).expanduser()
        sheet_name = parameters.get("sheet_name", 0)  # Default to first sheet

        if not file_path.exists():
            return self._error_response(f"File not found: {file_path}")

        try:
            # Load with pandas
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Compute statistics
            analysis = {
                "shape": {"rows": len(df), "columns": len(df.columns)},
                "columns": list(df.columns),
                "data_types": df.dtypes.astype(str).to_dict(),
                "missing_values": df.isnull().sum().to_dict(),
                "numeric_stats": {}
            }
            
            # Numeric column statistics
            numeric_cols = df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                analysis["numeric_stats"][col] = {
                    "mean": float(df[col].mean()),
                    "median": float(df[col].median()),
                    "std": float(df[col].std()),
                    "min": float(df[col].min()),
                    "max": float(df[col].max())
                }
            
            # Sample data
            analysis["sample_rows"] = df.head(5).to_dict(orient='records')
            
            return self._success_response(
                result=analysis,
                metadata={"analyzed_with": "pandas"}
            )
        
        except Exception as e:
            logger.error(f"Excel analysis error: {e}")
            return self._error_response(f"Analysis error: {e}")
    
    async def _format_excel(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Format Excel cells"""

        import openpyxl

        file_path = Path(parameters.get("file_path", "")).expanduser()
        sheet_name = parameters.get("sheet_name", "Sheet1")
        cell_range = parameters.get("range", "A1")
        formatting = parameters.get("formatting", {})

        if not file_path.exists():
            return self._error_response(f"File not found: {file_path}")

        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            
            if sheet_name not in wb.sheetnames:
                return self._error_response(f"Sheet '{sheet_name}' not found")
            
            ws = wb[sheet_name]
            
            # Apply formatting
            for cell in ws[cell_range]:
                if isinstance(cell, tuple):
                    for c in cell:
                        self._apply_cell_formatting(c, formatting)
                else:
                    self._apply_cell_formatting(cell, formatting)
            
            # Save
            wb.save(file_path)
            wb.close()
            
            return self._success_response(
                result={"formatted_range": cell_range},
                metadata={"formatting_applied": list(formatting.keys())}
            )
        
        except Exception as e:
            logger.error(f"Excel formatting error: {e}")
            return self._error_response(f"Formatting error: {e}")
    
    def _apply_cell_formatting(self, cell, formatting: Dict):
        """Apply formatting to a cell"""

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Font
        if "font" in formatting:
            font_opts = formatting["font"]
            cell.font = Font(
                name=font_opts.get("name", "Calibri"),
                size=font_opts.get("size", 11),
                bold=font_opts.get("bold", False),
                italic=font_opts.get("italic", False),
                color=font_opts.get("color")
            )

        # Fill
        if "fill" in formatting:
            fill_opts = formatting["fill"]
            cell.fill = PatternFill(
                start_color=fill_opts.get("color", "FFFFFF"),
                end_color=fill_opts.get("color", "FFFFFF"),
                fill_type="solid"
            )

        # Alignment
        if "alignment" in formatting:
            align_opts = formatting["alignment"]
            cell.alignment = Alignment(
                horizontal=align_opts.get("horizontal", "left"),
                vertical=align_opts.get("vertical", "top"),
                wrap_text=align_opts.get("wrap", False)
            )

        # Border
        if "border" in formatting:
            border_opts = formatting["border"]
            side_style = border_opts.get("style", "thin")
            cell.border = Border(
                left=Side(style=side_style),
                right=Side(style=side_style),
                top=Side(style=side_style),
                bottom=Side(style=side_style)
            )
    
    async def _add_chart(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Add chart to Excel file"""

        import openpyxl
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        file_path = Path(parameters.get("file_path", "")).expanduser()
        sheet_name = parameters.get("sheet_name", "Sheet1")
        chart_type = parameters.get("chart_type", "bar")
        data_range = parameters.get("range", "A1:B10")

        if not file_path.exists():
            return self._error_response(f"File not found: {file_path}")

        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path)

            if sheet_name not in wb.sheetnames:
                return self._error_response(f"Sheet '{sheet_name}' not found")

            ws = wb[sheet_name]

            # Create chart
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                return self._error_response(f"Unknown chart type: {chart_type}")

            # Set data
            data = Reference(ws, range_string=data_range)
            chart.add_data(data, titles_from_data=True)
            
            # Add chart to sheet
            ws.add_chart(chart, "E5")
            
            # Save
            wb.save(file_path)
            wb.close()
            
            return self._success_response(
                result={"chart_added": chart_type},
                metadata={"data_range": data_range}
            )
        
        except Exception as e:
            logger.error(f"Excel chart error: {e}")
            return self._error_response(f"Chart error: {e}")


# ============================================================================
# USAGE EXAMPLES
# ============================================================================

async def example_excel_operations():
    """Example Excel operations"""
    
    tool = ExcelProcessorTool()
    
    print("=" * 60)
    print("Excel Processor - Examples")
    print("=" * 60)
    
    # Example 1: Create Excel with data
    print("\n1. Create Excel File")
    result = await tool.execute({
        "action": "write",
        "file_path": "/tmp/sales_data.xlsx",
        "sheet_name": "Q4 Sales",
        "headers": ["Product", "Units", "Revenue"],
        "data": [
            ["Widget A", 150, 15000],
            ["Widget B", 230, 34500],
            ["Widget C", 180, 27000]
        ],
        "formulas": [
            {"cell": "D1", "formula": "=SUM(C2:C4)"}
        ]
    })
    print(f"Result: {result}")
    
    # Example 2: Read and analyze
    print("\n2. Analyze Excel File")
    result = await tool.execute({
        "action": "analyze",
        "file_path": "/tmp/sales_data.xlsx",
        "sheet_name": "Q4 Sales"
    })
    print(f"Analysis: {result}")
    
    # Example 3: Format cells
    print("\n3. Format Cells")
    result = await tool.execute({
        "action": "format",
        "file_path": "/tmp/sales_data.xlsx",
        "sheet_name": "Q4 Sales",
        "range": "A1:C1",
        "formatting": {
            "font": {"bold": True, "color": "FFFFFF"},
            "fill": {"color": "4472C4"},
            "alignment": {"horizontal": "center"}
        }
    })
    print(f"Result: {result}")


if __name__ == "__main__":
    if not OPENPYXL_AVAILABLE:
        print("❌ openpyxl not installed")
        print("\nInstall: pip install openpyxl xlsxwriter pandas")
    else:
        print("✅ Excel processor available")
        print("\nRun examples:")
        print("  python excel_processor.py")
